import openpyxl
from datetime import datetime
import os
from data import get_from_db
import glob
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

def get_img(user_id):
            files = []
            search_pattern = f"C:/engineer/image/photo_{user_id}_*.jpg"
            print(f"Ищу файлы по пути: {search_pattern}")
            files = glob.glob(search_pattern)
            return files if files else None

def create(user_id, caption=None):
    file_path = None
    data = get_from_db(user_id)['payload']
    print(data)
    if not data:
        return None
    if isinstance(data, dict):
        try:
            wb = openpyxl.load_workbook('Заказ - наряд.xlsx')
        except FileNotFoundError:
            print("Ошибка: Не найден файл шаблона 'Заказ - наряд.xlsx'")
            return None
            
        ws = wb.active
        ws.cell(8, 3, data['Вал сукноведущий'])

        photos = get_img(user_id)
        if photos:
            img_col = 67
            for img_path in photos:
                img = Image(img_path) 
                img.width = 200
                img.height = 300
                anchor = f"B{img_col}"
                ws.add_image(img, anchor)
                img_col += 18
            
        
        print(caption)
        if caption:
            ws.merge_cells('F71:L71')
            ws['F71'] = caption
            ws.cell(row=38, column=14).value = caption #убрать это
            ws.cell(row=38, column=14).alignment = openpyxl.styles.Alignment(wrapText=True)
            
        start_row = 14

        for row in range(start_row, 200):
            cell_name = ws.cell(row=row, column=2).value
                
            if not cell_name:
                continue

            work_name = str(cell_name).strip()

            if work_name in data:
                value = data[work_name]
            
                quantity = 0
                if value is True:
                    quantity = 1
                elif value is False:
                    quantity = 0
                elif isinstance(value, (int, float)):
                    quantity = value
                
                if quantity > 0:
                    ws.cell(row=row, column=12).value = quantity
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"Order_{user_id}_{timestamp}.xlsx"
                os.makedirs(r"c:\engineer\reports", exist_ok=True)
                file_path = os.path.join('reports', filename)
        wb.save(file_path)
        if photos:
            for i in photos:
                if os.path.exists(i):
                    os.remove(i)
        
        return file_path
    
def find_file(user_id):
    search_pattern = f"c:/engineer/reports/Order_{user_id}_*.xlsx"
    print(f"Ищу файлы по пути: {search_pattern}")
    files = glob.glob(search_pattern)
    
    if not files:
        return None
    last_file = max(files, key=os.path.getmtime)
    
    return last_file